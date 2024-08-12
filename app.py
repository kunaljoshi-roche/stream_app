import streamlit as st
import os
import pdfplumber as plumber
import PyPDF2 as pdf
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image
from datetime import datetime

# This must be the first Streamlit command used in your app, and it should go right here:
st.set_page_config(page_title="LCMS PDF Data Processor", layout="wide")
#adding one commit
# Custom CSS for styling (if you have added any)
st.markdown("""
    <style>
        .stButton>button {
            border-radius: 10px;
            border: 1px solid #4CAF50;
            background-color: #4CAF50;
            color: white;
        }
        .stButton>button:hover {
            background-color: #45a049;
        }
        .stTextInput>div>input {
            border-radius: 5px;
            padding: 10px;
            border: 1px solid #ccc;
        }
    </style>
""", unsafe_allow_html=True)

# Display the title and logo
col1, col2 = st.columns([9, 1])
with col1:
    st.title("LCMS PDF Data Processor")
with col2:
    logo = Image.open("Roche_Logo.png")
    st.image(logo, use_column_width=True)


st.markdown("---")

# Initialize the state for URL inputs
if "url_inputs" not in st.session_state:
    st.session_state.url_inputs = [""]

def add_url_input():
    if len(st.session_state.url_inputs) < 10:
        st.session_state.url_inputs.append("")

def remove_url_input(index):
    if len(st.session_state.url_inputs) > 1:
        st.session_state.url_inputs.pop(index)

# Function to render URL inputs dynamically
def render_url_inputs():
    for i in range(len(st.session_state.url_inputs)):
        with st.container():
            cols = st.columns([9, 0.5, 0.5])  # Adjusted column widths
            st.session_state.url_inputs[i] = cols[0].text_input(f"Enter PDF URL {i+1}:", value=st.session_state.url_inputs[i])
            if i > 0:
                cols[1].button("❌", key=f"remove_{i}", on_click=remove_url_input, args=(i,), help="Remove this URL")
            if i == len(st.session_state.url_inputs) - 1 and len(st.session_state.url_inputs) < 10:
                cols[2].button("➕", key=f"add_{i}", on_click=add_url_input, help="Add another URL")

# Render URL input fields
render_url_inputs()

# Filter out empty inputs
urls = [url.strip() for url in st.session_state.url_inputs if url.strip()]

def extract_data_from_pdf(file_path):
    data = pdf.PdfReader(file_path)
    Obs_mass, sam_pos, flp = [], [], []
    
    for i in range(len(data.pages)):
        page_no = data.pages[i]
        txt = page_no.extract_text().split()
        
        for j, x in enumerate(txt):
            if x == "(Da)":
                ind_1 = j + 10
                ind_temp = ind_1 + 5
                if len(txt[ind_temp]) == 9:
                    Obs_mass.append([txt[ind_1], txt[ind_temp]])
                else:
                    Obs_mass.append(txt[ind_1])
                ind_2 = j - 8
                sam_pos.append(txt[ind_2])
                ind_3 = j + 12
                flp.append(txt[ind_3])
                
    return sam_pos, Obs_mass, flp

def merge_and_clean_data(sam_pos, Obs_mass, flp, file_path):
    df = pd.DataFrame({
        "Sample Position": sam_pos,
        "Observed mass (Da)": Obs_mass,
        "FLP UV % Area": flp
    })
    
    position = []
    with plumber.open(file_path) as pdf:
        for i in range(len(pdf.pages)):
            try:
                table = pdf.pages[i].extract_table()
                if table and table[0][2] == 'Sample type':
                    position.append(table[1][-2])
            except (IndexError, TypeError):
                continue

    sample_df = pd.DataFrame({"Sample Position": position})

    df["Sample Position"] = df["Sample Position"].apply(lambda x: x[:-1])
    df_new = pd.merge(df, sample_df, on="Sample Position", how='outer')
    df_new.drop(df_new.loc[df_new["Sample Position"].str.len() < 5].index, inplace=True)

    return df_new

def custom_sort_logic(value):
    match = re.match(r'(\d+)\s*:\s*(\w+)\s*,\s*(\d+)', value)
    if match:
        section1 = int(match.group(1))
        section2 = match.group(2)
        section3 = int(match.group(3))
        return (section1, section3, section2)
    else:
        return (float('inf'), '', float('inf'))

def sort_dataframe(df, column_name):
    df_sorted = df.sort_values(by=column_name, key=lambda x: x.apply(custom_sort_logic))
    return df_sorted

def expand_observed_mass(df):
    expanded_rows = []
    for idx, row in df.iterrows():
        if isinstance(row["Observed mass (Da)"], list):
            for i, obs_mass in enumerate(row["Observed mass (Da)"]):
                if pd.notna(obs_mass):
                    obs_mass = round(float(obs_mass))  # Round to the nearest integer
                if i == 0:
                    expanded_rows.append([row["Sample Position"], obs_mass, row["FLP UV % Area"]])
                else:
                    expanded_rows.append([row["Sample Position"], obs_mass, ''])
        else:
            if pd.notna(row["Observed mass (Da)"]):
                obs_mass = round(float(row["Observed mass (Da)"]))  # Round to the nearest integer
            else:
                obs_mass = row["Observed mass (Da)"]  # Keep as is if NaN
            expanded_rows.append([row["Sample Position"], obs_mass, row["FLP UV % Area"]])
    expanded_df = pd.DataFrame(expanded_rows, columns=["Sample Position", "Observed mass (Da)", "FLP UV % Area"])
    return expanded_df

def clean_sample_position(df):
    df["Sample Position"] = df["Sample Position"].apply(lambda x: re.sub(r'(\w+)\s*,\s*(\d+)', r'\1,\2', x))
    return df

def save_to_excel(df, file_path):
    timestamp = datetime.now().strftime("%Y/%m/%d_%H:%M")
    excel_path = os.path.join(os.path.dirname(file_path), f"{os.path.basename(file_path).replace('.pdf', '')}_Processed_{timestamp}.xlsx")
    df.to_excel(excel_path, index=False, engine='openpyxl')
    
    wb = load_workbook(excel_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    positions_with_multiple_obs_mass = df[df.duplicated('Sample Position', keep=False)]['Sample Position'].unique()
    
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):  # Start from the second row
        current_position = row[0].value
        if current_position in positions_with_multiple_obs_mass:
            row[1].fill = yellow_fill

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):  # Start from the second row and second column
        if all(cell.value in (None, '') for cell in row):
            for cell in row:
                cell.fill = red_fill
    
    wb.save(excel_path)
    return excel_path

if st.button('Process PDF Files'):
    if urls:
        processed_files = []
        for url in urls:
            file_path = url
            if os.path.isfile(file_path):
                sam_pos, Obs_mass, flp = extract_data_from_pdf(file_path)
                df_new = merge_and_clean_data(sam_pos, Obs_mass, flp, file_path)
                df_new = clean_sample_position(df_new)
                sorted_df = sort_dataframe(df_new, 'Sample Position')
                expanded_df = expand_observed_mass(sorted_df)
                processed_file = save_to_excel(expanded_df, file_path)
                processed_files.append(processed_file)
                st.success(f"File processed successfully: {os.path.basename(processed_file)}")
                with open(processed_file, "rb") as f:
                    st.download_button(f"Download {os.path.basename(processed_file)}", f, file_name=os.path.basename(processed_file))
            else:
                st.error(f"Invalid file path: {file_path}")
    else:
        st.error("Please enter at least one valid file path.")